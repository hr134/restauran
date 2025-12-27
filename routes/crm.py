from flask import Blueprint, render_template, redirect, url_for, flash, request, make_response
from models.models import db, User, Order
from sqlalchemy import func

crm_bp = Blueprint('crm', __name__, url_prefix='/crm')

def admin_required(f):
    """Decorator to require admin or manager role"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        from flask import session
        if session.get('role') not in ['admin', 'manager']:
            flash('Access denied. Manager privileges required.', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated_function

@crm_bp.route('/')
@admin_required
def index():
    """Customer management dashboard"""
    # Get all customers (users with role='customer')
    customers = User.query.filter_by(role='customer').all()
    
    # Calculate customer statistics
    customer_stats = []
    for customer in customers:
        total_orders = Order.query.filter_by(user_id=customer.id).count()
        total_spent = db.session.query(func.sum(Order.total)).filter_by(user_id=customer.id).scalar() or 0
        avg_order_value = total_spent / total_orders if total_orders > 0 else 0
        
        # Determine customer tier
        if total_spent > 10000:
            tier = 'VIP'
        elif total_spent > 5000:
            tier = 'Gold'
        elif total_spent > 1000:
            tier = 'Silver'
        else:
            tier = 'Regular'
        
        customer_stats.append({
            'customer': customer,
            'total_orders': total_orders,
            'total_spent': total_spent,
            'avg_order_value': avg_order_value,
            'tier': tier
        })
    
    # Sort by total spent (descending)
    customer_stats.sort(key=lambda x: x['total_spent'], reverse=True)
    
    from flask import session
    template_folder = 'admin' if session.get('role') == 'admin' else 'manager'
    return render_template(f'{template_folder}/customer_management.html', customer_stats=customer_stats)

@crm_bp.route('/view/<int:id>')
@admin_required
def view_customer(id):
    """View customer details and purchase history"""
    customer = User.query.get_or_404(id)
    
    if customer.role != 'customer':
        flash('This user is not a customer', 'warning')
        return redirect(url_for('crm.index'))
    
    # Get customer's orders
    orders = Order.query.filter_by(user_id=customer.id).order_by(Order.created_at.desc()).all()
    
    # Parse items for template
    import json
    for order in orders:
        try:
            order.items_parsed = json.loads(order.items)
        except:
            order.items_parsed = []
    
    # Calculate statistics
    total_orders = len(orders)
    total_spent = sum(order.total for order in orders)
    avg_order_value = total_spent / total_orders if total_orders > 0 else 0
    
    # Get favorite items (most ordered)
    from models.models import SaleItem, MenuItem
    favorite_items = db.session.query(
        MenuItem.name,
        func.sum(SaleItem.quantity).label('total_quantity')
    ).join(SaleItem).join(Order).filter(
        Order.user_id == customer.id
    ).group_by(MenuItem.id).order_by(func.sum(SaleItem.quantity).desc()).limit(5).all()
    
    from flask import session
    template_folder = 'admin' if session.get('role') == 'admin' else 'manager'
    return render_template(f'{template_folder}/customer_detail.html',
                         customer=customer,
                         orders=orders,
                         total_orders=total_orders,
                         total_spent=total_spent,
                         avg_order_value=avg_order_value,
                         favorite_items=favorite_items)

@crm_bp.route('/export-customers')
@admin_required
def export_customers():
    """Export customer data to Excel"""
    from openpyxl import Workbook
    from io import BytesIO
    from flask import make_response
    from datetime import datetime
    
    customers = User.query.filter_by(role='customer').all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Header
    ws.append(['Username', 'Full Name', 'Email', 'Phone', 'Total Orders', 'Total Spent', 'Member Since'])
    
    # Styling (optional but nice for Excel)
    from openpyxl.styles import Font
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for customer in customers:
        total_orders = Order.query.filter_by(user_id=customer.id).count()
        total_spent = db.session.query(func.sum(Order.total)).filter_by(user_id=customer.id).scalar() or 0
        
        ws.append([
            customer.username,
            customer.full_name or 'N/A',
            customer.email or 'N/A',
            customer.phone or 'N/A',
            total_orders,
            total_spent,
            customer.created_at.strftime('%Y-%m-%d') if customer.created_at else 'N/A'
        ])
    
    # Format Currency column (Total Spent is column F/6)
    for row in range(2, len(customers) + 2):
        ws.cell(row=row, column=6).number_format = '"৳"#,##0.00'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers["Content-Disposition"] = f"attachment; filename=customers_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

@crm_bp.route('/order/<int:order_id>')
@admin_required
def get_order(order_id):
    """Get order details for manager/admin (AJAX)"""
    from flask import jsonify
    import json
    order = Order.query.get_or_404(order_id)
    
    try:
        items = json.loads(order.items)
    except:
        items = []
        
    user = User.query.get(order.user_id)
    return jsonify({
        'id': order.id,
        'unique_order_number': order.unique_order_number,
        'total': order.total,
        'status': order.status,
        'created_at': order.created_at.strftime('%Y-%m-%d %H:%M'),
        'customer': user.full_name if user else 'Unknown',
        'username': user.username if user else 'anon',
        'phone': order.phone,
        'address_street': order.address_street,
        'address_city': order.address_city,
        'address_district': order.address_district,
        'items': items
    })
